import json
import csv
import io
import re
import time
import traceback

from django.shortcuts import render
from django.http import JsonResponse, StreamingHttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.core.mail import EmailMessage
from django.conf import settings


def home(request):
    """Render the main SkyMailer interface."""
    return render(request, 'email_form.html', {
        'sender_email': settings.EMAIL_HOST_USER,
    })


@require_POST
def upload_csv(request):
    """Parse an uploaded CSV/TXT/XLSX file and return the list of recipients with variables."""
    csv_file = request.FILES.get('file')
    sheet_name = request.POST.get('sheet_name', '')
    
    if not csv_file:
        return JsonResponse({'error': 'No se proporcionó archivo.'}, status=400)

    try:
        file_content = csv_file.read()
        file_ext = csv_file.name.lower().rsplit('.', 1)[-1] if '.' in csv_file.name else ''
        
        # Handle Excel files
        if file_ext in ('xlsx', 'xls'):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
            sheets = wb.sheetnames
            
            # If no sheet specified and multiple sheets, return sheet list for user to choose
            if not sheet_name and len(sheets) > 1:
                wb.close()
                return JsonResponse({
                    'sheets': sheets,
                    'needs_sheet_selection': True,
                })
            
            # Use specified sheet or first/only sheet
            target_sheet = sheet_name if sheet_name else sheets[0]
            if target_sheet not in sheets:
                wb.close()
                return JsonResponse({'error': f'La hoja "{target_sheet}" no existe.'}, status=400)
            
            ws = wb[target_sheet]
            
            # Convert sheet to CSV-like text
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
            
            if not rows:
                return JsonResponse({'error': 'La hoja está vacía.'}, status=400)
            
            # First row = headers
            headers = [str(h).strip().lower() if h is not None else f'col_{i}' for i, h in enumerate(rows[0])]
            
            # Build text as CSV
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(headers)
            for row in rows[1:]:
                writer.writerow([str(c) if c is not None else '' for c in row])
            text = output.getvalue()
            delimiter = ','
        else:
            # CSV/TXT file
            try:
                text = file_content.decode('utf-8-sig')
            except UnicodeDecodeError:
                text = file_content.decode('latin-1')

            # Detect delimiter
            first_line = text.split('\n')[0]
            if '\t' in first_line:
                delimiter = '\t'
            elif ';' in first_line:
                delimiter = ';'
            else:
                delimiter = ','

        reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
        
        # Normalize field names
        if reader.fieldnames is None:
            return JsonResponse({'error': 'El archivo está vacío.'}, status=400)
        
        reader.fieldnames = [f.strip().lower() for f in reader.fieldnames]

        
        recipients = []
        variables = list(reader.fieldnames)
        
        # Find the email column
        email_col = None
        email_names = ('email', 'correo', 'mail', 'e-mail', 'email address', 'dirección',
                       'email_destinatario', 'destinatario', 'to', 'recipient')
        for col in reader.fieldnames:
            if col in email_names:
                email_col = col
                break
        
        # Fallback: find column whose name contains 'email' or 'mail'
        if email_col is None:
            for col in reader.fieldnames:
                if 'email' in col or 'mail' in col or 'correo' in col:
                    email_col = col
                    break
        
        if email_col is None:
            # Try to find a column that looks like emails by content
            sample = text.split('\n')
            if len(sample) > 1:
                for col in reader.fieldnames:
                    reader2 = csv.DictReader(io.StringIO(text), delimiter=delimiter)
                    reader2.fieldnames = [f.strip().lower() for f in reader2.fieldnames]
                    for row in reader2:
                        val = row.get(col, '').strip()
                        if '@' in val and '.' in val:
                            email_col = col
                            break
                        break
            
            if email_col is None:
                return JsonResponse({
                    'error': 'No se encontró una columna de email. Asegúrate de que tu archivo tenga una columna con "email" en el nombre.',
                }, status=400)
        
        # Detect batch mode columns (per-row subject and message)
        subject_col = None
        message_col = None
        
        # Exact matches first
        subject_names = ('asunto', 'subject', 'tema')
        message_names = ('mensaje', 'message', 'cuerpo', 'body', 'contenido')
        
        for col in reader.fieldnames:
            if col in subject_names:
                subject_col = col
            if col in message_names:
                message_col = col
        
        # Fuzzy match: column name CONTAINS 'subject', 'asunto', 'body', 'mensaje', etc.
        if subject_col is None:
            for col in reader.fieldnames:
                if 'subject' in col or 'asunto' in col or 'tema' in col:
                    subject_col = col
                    break
        
        if message_col is None:
            for col in reader.fieldnames:
                if 'body' in col or 'mensaje' in col or 'cuerpo' in col or 'contenido' in col:
                    message_col = col
                    break
        
        batch_mode = subject_col is not None and message_col is not None
        
        for row in reader:
            email = row.get(email_col, '').strip()
            if email and '@' in email:
                recipient_data = {'email': email}
                for var in variables:
                    if var != email_col:
                        recipient_data[var] = row.get(var, '').strip()
                # In batch mode, store subject and message with special keys
                if batch_mode:
                    recipient_data['_asunto'] = row.get(subject_col, '').strip()
                    recipient_data['_mensaje'] = row.get(message_col, '').strip()
                recipients.append(recipient_data)
        
        # Remove the email column and batch columns from display variables
        hide_cols = {email_col, subject_col, message_col} - {None}
        display_vars = [v for v in variables if v not in hide_cols]
        
        return JsonResponse({
            'recipients': recipients,
            'variables': display_vars,
            'count': len(recipients),
            'email_column': email_col,
            'batch_mode': batch_mode,
        })
    except Exception as e:
        return JsonResponse({'error': f'Error al procesar el archivo: {str(e)}'}, status=400)


@require_POST
def send_emails_ajax(request):
    """Send emails one by one with rate limiting, returning results as a JSON stream."""
    try:
        subject_template = request.POST.get('asunto', '')
        message_template = request.POST.get('mensaje', '')
        recipients_json = request.POST.get('destinatarios', '[]')
        is_html = request.POST.get('is_html', 'true') == 'true'
        batch_mode = request.POST.get('batch_mode', 'false') == 'true'
        delay_seconds = float(request.POST.get('delay', '3'))
        
        # Clamp delay between 2 and 10 seconds
        delay_seconds = max(2, min(10, delay_seconds))
        
        recipients = json.loads(recipients_json)
        
        if not recipients:
            return JsonResponse({'error': 'No hay destinatarios.'}, status=400)
        
        # In batch mode each recipient has its own subject; in normal mode require global subject
        if not batch_mode and not subject_template.strip():
            return JsonResponse({'error': 'El asunto no puede estar vacío.'}, status=400)
        
        # Collect attachments
        attachments = []
        for key in request.FILES:
            if key.startswith('attachment'):
                f = request.FILES[key]
                attachments.append({
                    'filename': f.name,
                    'content': f.read(),
                    'mimetype': f.content_type,
                })

        def generate():
            """Stream results as newline-delimited JSON."""
            results = []
            for i, recipient in enumerate(recipients):
                email_addr = recipient.get('email', '').strip()
                if not email_addr:
                    results.append({'index': i, 'email': email_addr, 'status': 'error', 'message': 'Email vacío'})
                    yield json.dumps(results[-1]) + '\n'
                    continue
                
                # In batch mode, use per-recipient subject and message
                if batch_mode:
                    subject = recipient.get('_asunto', subject_template)
                    message = recipient.get('_mensaje', message_template)
                else:
                    subject = subject_template
                    message = message_template
                
                # Replace variables in subject and message
                for key, value in recipient.items():
                    if key.startswith('_'):  # Skip internal fields
                        continue
                    placeholder = '{' + key + '}'
                    subject = subject.replace(placeholder, str(value))
                    message = message.replace(placeholder, str(value))
                
                try:
                    email = EmailMessage(
                        subject=subject,
                        body=message,
                        from_email=settings.EMAIL_HOST_USER,
                        to=[email_addr],
                    )
                    
                    if is_html:
                        email.content_subtype = 'html'
                    
                    # Add attachments
                    for att in attachments:
                        email.attach(att['filename'], att['content'], att['mimetype'])
                    
                    email.send(fail_silently=False)
                    
                    results.append({
                        'index': i,
                        'email': email_addr,
                        'status': 'success',
                        'message': 'Enviado correctamente',
                    })
                except Exception as e:
                    results.append({
                        'index': i,
                        'email': email_addr,
                        'status': 'error',
                        'message': str(e),
                    })
                
                yield json.dumps(results[-1]) + '\n'
                
                # Rate limiting - wait between emails to avoid Gmail blocking
                if i < len(recipients) - 1:
                    time.sleep(delay_seconds)
        
        response = StreamingHttpResponse(generate(), content_type='application/x-ndjson')
        response['X-Accel-Buffering'] = 'no'
        response['Cache-Control'] = 'no-cache'
        return response
    
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Formato de destinatarios inválido.'}, status=400)
    except Exception as e:
        return JsonResponse({'error': f'Error del servidor: {str(e)}'}, status=500)


def get_templates(request):
    """Return predefined email templates."""
    templates = [
        {
            'id': 'newsletter',
            'name': '📰 Newsletter',
            'subject': 'Novedades de {empresa} — {mes}',
            'body': '''<div style="font-family: 'Segoe UI', Arial, sans-serif; max-width: 600px; margin: 0 auto;">
    <h2 style="color: #6C63FF;">Hola {nombre},</h2>
    <p>Te traemos las últimas novedades de este mes:</p>
    <ul>
        <li>📌 Novedad 1</li>
        <li>📌 Novedad 2</li>
        <li>📌 Novedad 3</li>
    </ul>
    <p>¡Gracias por seguirnos!</p>
    <p style="color: #888; font-size: 0.9em;">— El equipo de {empresa}</p>
</div>''',
        },
        {
            'id': 'promo',
            'name': '🎯 Promoción',
            'subject': '¡Oferta exclusiva para ti, {nombre}!',
            'body': '''<div style="font-family: 'Segoe UI', Arial, sans-serif; max-width: 600px; margin: 0 auto; text-align: center;">
    <h1 style="color: #FF6B6B;">🔥 ¡Oferta Especial!</h1>
    <p style="font-size: 1.2em;">Hola <strong>{nombre}</strong>,</p>
    <p>Tenemos una oferta exclusiva preparada solo para ti.</p>
    <div style="background: linear-gradient(135deg, #6C63FF, #FF6B6B); color: white; padding: 20px; border-radius: 12px; margin: 20px 0;">
        <h2 style="margin: 0; color: white;">20% de descuento</h2>
        <p style="margin: 5px 0 0;">Usa el código: <strong>PROMO2026</strong></p>
    </div>
    <p>¡No te lo pierdas!</p>
    <p style="color: #888;">— {empresa}</p>
</div>''',
        },
        {
            'id': 'notification',
            'name': '🔔 Notificación',
            'subject': 'Información importante — {empresa}',
            'body': '''<div style="font-family: 'Segoe UI', Arial, sans-serif; max-width: 600px; margin: 0 auto;">
    <h2>Hola {nombre},</h2>
    <p>Queríamos informarte sobre lo siguiente:</p>
    <div style="background: #f0f0f0; padding: 15px; border-left: 4px solid #6C63FF; border-radius: 4px; margin: 15px 0;">
        <p style="margin: 0;">Tu mensaje de notificación aquí.</p>
    </div>
    <p>Si tienes alguna pregunta, no dudes en contactarnos.</p>
    <p>Saludos,<br><strong>{empresa}</strong></p>
</div>''',
        },
        {
            'id': 'welcome',
            'name': '👋 Bienvenida',
            'subject': '¡Bienvenido/a a {empresa}, {nombre}!',
            'body': '''<div style="font-family: 'Segoe UI', Arial, sans-serif; max-width: 600px; margin: 0 auto; text-align: center;">
    <h1 style="color: #6C63FF;">¡Bienvenido/a! 🎉</h1>
    <p style="font-size: 1.2em;">Hola <strong>{nombre}</strong>,</p>
    <p>Nos alegra mucho que formes parte de <strong>{empresa}</strong>.</p>
    <p>Estamos aquí para ayudarte en lo que necesites. No dudes en escribirnos.</p>
    <div style="margin-top: 30px;">
        <p style="color: #888; font-size: 0.9em;">— El equipo de {empresa}</p>
    </div>
</div>''',
        },
        {
            'id': 'blank',
            'name': '📝 En blanco',
            'subject': '',
            'body': '''<div style="font-family: 'Segoe UI', Arial, sans-serif; max-width: 600px; margin: 0 auto;">
    <p>Hola {nombre},</p>
    <p></p>
    <p>Saludos,<br>{empresa}</p>
</div>''',
        },
    ]
    return JsonResponse({'templates': templates})