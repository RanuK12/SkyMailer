"""
Script to reformat all email bodies in the Excel file to professional HTML.
Uses Import_Instantly sheet for metadata cross-referenced by email.
Tone: Cordobés, personal, humano — no "europeo que baja línea"
sino "cordobés que hizo carrera afuera y vuelve a aportar".
"""
import openpyxl
import re

INPUT_FILE = r'C:\Users\emilio\Desktop\Oficina Ranuk\Ranuk-Outreach\emails_listos_todos.xlsx'
OUTPUT_FILE = r'C:\Users\emilio\Desktop\Oficina Ranuk\Ranuk-Outreach\emails_listos_todos.xlsx'

print("Loading workbook...")
wb = openpyxl.load_workbook(INPUT_FILE)

# ── Build metadata lookup from Import_Instantly ──
ws_meta = wb['Import_Instantly']
meta_rows = list(ws_meta.iter_rows(values_only=True))
meta_headers = [str(h).strip() for h in meta_rows[0]]

leads = {}
for row in meta_rows[1:]:
    data = dict(zip(meta_headers, row))
    email = str(data.get('Email', '')).strip().lower()
    if email:
        leads[email] = {
            'empresa': str(data.get('Nombre_Empresa', '')),
            'sitio': str(data.get('Sitio_Empresa', '')),
            'rubro': str(data.get('Rubro', '')),
            'falencia': str(data.get('Falencia_Tecnologica', '')),
            'score': str(data.get('Score', '')),
        }

print(f"  {len(leads)} leads loaded")


# ══════════════════════════════════════════════
#  HTML TEMPLATES
# ══════════════════════════════════════════════

def format_cold_email(original_body, empresa, meta):
    """Email 1 - Cold: tono cordobés, personal, humano."""
    falencia = meta.get('falencia', '')
    sitio = meta.get('sitio', '')
    
    # Extract specific technical issues from the → line
    issues = []
    for line in original_body.split('\n'):
        if '\u2192' in line:
            # Skip calendly/ranuk links
            if 'calendly' in line.lower() or 'ranuk' in line.lower():
                continue
            parts = line.strip().lstrip('\u2192').strip().split('\u00b7')
            for p in parts:
                p = p.strip()
                if p and len(p) > 5:
                    issues.append(p)
    
    if not issues:
        issues = [falencia.capitalize()] if falencia and falencia != 'nan' else ['Oportunidades de mejora']
    
    issues_html = ''.join(f'<li style="padding:5px 0;color:#444">{iss}</li>' for iss in issues[:3])
    
    calendly = 'https://calendly.com/emilio-ranuk/30min'

    html = f'''<div style="font-family:'Segoe UI',Arial,sans-serif;max-width:600px;margin:0 auto;color:#2d2d2d;line-height:1.75;font-size:15px">

<p>Hola equipo de {empresa},</p>

<p>Soy Emilio, cordobés e ingeniero en sistemas. Hice gran parte de mi carrera afuera \u2014 laburé en la infraestructura de pricing de <strong>Booking.com</strong> en \u00c1msterdam y en proyectos de datos para clientes Fortune 500 en <strong>Accenture</strong> en Roma. Me form\u00e9 en entornos donde la vara t\u00e9cnica es altísima.</p>

<p>Ahora estoy de vuelta en C\u00f3rdoba con algo claro: <strong>traer ese nivel de tecnolog\u00eda a empresas de ac\u00e1</strong>, de forma directa y sin intermediarios.</p>

<p>Antes de escribirles, revis\u00e9 <a href="{sitio}" style="color:#6C63FF;text-decoration:none;border-bottom:1px solid #6C63FF">{sitio}</a> y detect\u00e9 algunas cosas concretas que se pueden mejorar:</p>

<div style="background:#f8f5ff;border-left:4px solid #6C63FF;padding:16px 20px;border-radius:0 8px 8px 0;margin:18px 0">
<ul style="margin:0;padding-left:18px">
{issues_html}
</ul>
</div>

<p>No lo digo como cr\u00edtica, al contrario \u2014 son el tipo de cosas que en las empresas donde trabaj\u00e9 se resuelven en <strong>2-3 semanas</strong>, y que cuando se atacan bien liberan <strong>15-30% de capacidad operativa</strong>. En Argentina casi nadie las mira porque no hay costumbre, pero la diferencia se nota r\u00e1pido.</p>

<table style="width:100%;border-collapse:separate;border-spacing:8px;margin:18px 0">
<tr>
<td style="padding:14px 16px;background:#fafafa;border-radius:10px;vertical-align:top;width:33%">
<strong style="color:#6C63FF">&#9997;&#65039; Trabajo directo</strong><br>
<span style="font-size:13px;color:#666">Escribo el c\u00f3digo yo, con mi estudio. Nada de subcontratar.</span>
</td>
<td style="padding:14px 16px;background:#fafafa;border-radius:10px;vertical-align:top;width:33%">
<strong style="color:#6C63FF">&#127919; Est\u00e1ndares reales</strong><br>
<span style="font-size:13px;color:#666">CI/CD, tests, docs. Lo que aprend\u00ed afuera, aplicado ac\u00e1.</span>
</td>
<td style="padding:14px 16px;background:#fafafa;border-radius:10px;vertical-align:top;width:33%">
<strong style="color:#6C63FF">&#128230; Proyectos acotados</strong><br>
<span style="font-size:13px;color:#666">8-12 semanas, entregables concretos. Sin retainers eternos.</span>
</td>
</tr>
</table>

<p><strong>\u00bfTendr\u00edas 30 minutos la pr\u00f3xima semana para charlar?</strong> Sin venta, sin propuesta armada \u2014 simplemente te cuento qu\u00e9 ver\u00eda yo en {empresa} si fuera tu CTO por un rato. Si no tiene sentido, lo dejamos ah\u00ed y ya.</p>

<p style="text-align:center;margin:24px 0">
<a href="{calendly}" style="display:inline-block;background:linear-gradient(135deg,#6C63FF,#a855f7);color:white;padding:14px 32px;border-radius:8px;text-decoration:none;font-weight:600;font-size:15px;box-shadow:0 4px 15px rgba(108,99,255,0.25)">&#128197; Agendar llamada de 30 min</a>
</p>

<p style="font-size:13px;color:#888">O respond\u00e9 este mail si prefer\u00eds \u2014 te contesto en el d\u00eda.</p>

<hr style="border:none;border-top:1px solid #eee;margin:28px 0">

<table style="width:100%"><tr>
<td style="vertical-align:top;padding-right:16px">
<p style="margin:0;font-weight:700;font-size:15px">Emilio Ranucoli</p>
<p style="margin:3px 0;font-size:12px;color:#888">Ingeniero en Sistemas \u00b7 Fundador, Ranuk IT Solutions</p>
<p style="margin:3px 0;font-size:12px;color:#888">Ex-Booking.com \u00b7 Ex-Accenture \u00b7 AWS Certified</p>
<p style="margin:6px 0 0"><a href="https://ranuk.dev/ranuk-it" style="color:#6C63FF;font-size:13px;text-decoration:none">ranuk.dev/ranuk-it</a></p>
</td>
</tr></table>

<p style="font-size:11px;color:#bbb;margin-top:24px;border-top:1px solid #f0f0f0;padding-top:12px">Si prefer\u00eds no recibir m\u00e1s correos de mi parte, respond\u00e9 \u201cremover\u201d y sal\u00eds de la lista en el d\u00eda. No volvemos a escribir.</p>

</div>'''
    return html


def format_followup_email(original_body, empresa, meta):
    """Email 2 - Follow-Up: directo, cálido, sin presión."""
    falencia = meta.get('falencia', '')
    sitio = meta.get('sitio', '')
    if falencia == 'nan':
        falencia = 'las mejoras técnicas'

    html = f'''<div style="font-family:'Segoe UI',Arial,sans-serif;max-width:600px;margin:0 auto;color:#2d2d2d;line-height:1.75;font-size:15px">

<p>Emilio de nuevo \u2014</p>

<p>La semana pasada les dej\u00e9 una nota sobre <strong>{falencia}</strong> que detect\u00e9 en <a href="{sitio}" style="color:#6C63FF;text-decoration:none;border-bottom:1px solid #6C63FF">{sitio}</a>. S\u00e9 que estos mails a veces quedan enterrados en la bandeja, as\u00ed que voy al punto:</p>

<div style="background:#f8f5ff;border-left:4px solid #6C63FF;padding:20px 24px;border-radius:0 10px 10px 0;margin:20px 0">
<p style="margin:0;font-weight:600;color:#333;font-size:16px">Una sola pregunta:</p>
<p style="margin:10px 0 0;font-size:15px">\u00bfEste tema est\u00e1 en el radar de <strong>{empresa}</strong> para los pr\u00f3ximos meses, o por ahora no les mueve la aguja?</p>
</div>

<table style="width:100%;border-collapse:separate;border-spacing:10px;margin:20px 0">
<tr>
<td style="padding:18px 20px;background:#fff5f5;border-radius:10px;border:1px solid #fee;vertical-align:top;width:48%">
<p style="margin:0;font-weight:700;color:#e74c3c">\u274c \u201cNo es prioridad\u201d</p>
<p style="margin:8px 0 0;font-size:13px;color:#666">Te saco de la lista y listo. Cero drama, cero insistencia.</p>
</td>
<td style="padding:18px 20px;background:#f0fff4;border-radius:10px;border:1px solid #d4edda;vertical-align:top;width:48%">
<p style="margin:0;font-weight:700;color:#27ae60">\u2705 \u201cS\u00ed, pero es complejo\u201d</p>
<p style="margin:8px 0 0;font-size:13px;color:#666">Te preparo un <strong>mini audit gratis de 1 p\u00e1gina</strong> para que llegues a la charla con datos, no con un pitch.</p>
</td>
</tr>
</table>

<p>Respond\u00e9 lo que sea, con una sola palabra alcanza.</p>

<p>Un abrazo,<br><strong>Emilio</strong></p>

</div>'''
    return html


def format_breakup_email(original_body, empresa, meta):
    """Email 3 - Breakup: generoso, sin rencor, puerta abierta."""
    falencia = meta.get('falencia', '')
    if falencia == 'nan':
        falencia = 'mejoras técnicas'

    html = f'''<div style="font-family:'Segoe UI',Arial,sans-serif;max-width:600px;margin:0 auto;color:#2d2d2d;line-height:1.75;font-size:15px">

<p>\u00daltima nota y te dejo tranquilo \u2014</p>

<p>Les comparto un recurso que puede servirles aunque no terminemos trabajando juntos:</p>

<div style="background:linear-gradient(135deg,#f8f5ff,#fff5f8);padding:24px;border-radius:14px;margin:20px 0;text-align:center;border:1px solid rgba(108,99,255,0.1)">
<p style="margin:0">
<a href="https://ranuk.dev/ranuk-it/" style="color:#6C63FF;font-weight:700;font-size:16px;text-decoration:none">&#128279; ranuk.dev/ranuk-it</a>
</p>
<p style="margin:10px 0 0;font-size:13px;color:#666;font-style:italic">\u201cC\u00f3mo un ingeniero cordob\u00e9s que laburaba en Booking.com ahora construye software para empresas de ac\u00e1 \u2014 sin intermediarios\u201d</p>
</div>

<p>Si en alg\u00fan momento <strong>{falencia}</strong> pasa a ser prioridad, saben d\u00f3nde encontrarme. Los dejo fuera de la secuencia \u2014 <strong>no van a recibir m\u00e1s correos m\u00edos.</strong></p>

<hr style="border:none;border-top:1px solid #eee;margin:28px 0">

<p style="margin:0">Gracias por el tiempo de leer,</p>
<p style="margin:6px 0 0;font-weight:700">Emilio Ranucoli</p>
<p style="margin:3px 0"><a href="https://ranuk.dev/ranuk-it" style="color:#6C63FF;font-size:13px;text-decoration:none">ranuk.dev/ranuk-it</a></p>

</div>'''
    return html


# ══════════════════════════════════════════════
#  MAIN PROCESSING
# ══════════════════════════════════════════════

sheet_configs = {
    'Email_1_Cold': {'body_col': 'email_1_body', 'email_col': 'email_destinatario', 'empresa_col': 'empresa', 'formatter': format_cold_email},
    'Email_2_FollowUp': {'body_col': 'email_2_body', 'email_col': 'email_destinatario', 'empresa_col': 'empresa', 'formatter': format_followup_email},
    'Email_3_Breakup': {'body_col': 'email_3_body', 'email_col': 'email_destinatario', 'empresa_col': 'empresa', 'formatter': format_breakup_email},
}

for sheet_name, config in sheet_configs.items():
    if sheet_name not in wb.sheetnames:
        print(f"Sheet {sheet_name} not found, skipping")
        continue
    
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h).strip().lower() if h else '' for h in rows[0]]
    
    body_idx = email_idx = empresa_idx = None
    for i, h in enumerate(headers):
        if h == config['body_col']: body_idx = i
        if h == config['empresa_col']: empresa_idx = i
        if h == config['email_col']: email_idx = i
    
    if body_idx is None:
        print(f"  Body column not found in {sheet_name}")
        continue
    
    print(f"Processing {sheet_name}: {len(rows)-1} emails...")
    
    updated = 0
    for row_idx in range(2, len(rows) + 1):
        cell = ws.cell(row=row_idx, column=body_idx + 1)
        empresa_cell = ws.cell(row=row_idx, column=empresa_idx + 1) if empresa_idx is not None else None
        email_cell = ws.cell(row=row_idx, column=email_idx + 1) if email_idx is not None else None
        
        original_body = str(cell.value) if cell.value else ''
        empresa = str(empresa_cell.value) if empresa_cell and empresa_cell.value else ''
        email_addr = str(email_cell.value).strip().lower() if email_cell and email_cell.value else ''
        
        meta = leads.get(email_addr, {'falencia': '', 'sitio': '', 'rubro': ''})
        
        if original_body.strip():
            html_body = config['formatter'](original_body, empresa, meta)
            cell.value = html_body
            updated += 1
    
    print(f"  OK - {sheet_name}: {updated} emails")

print(f"\nSaving...")
wb.save(OUTPUT_FILE)
wb.close()
print("Done! All emails reformatted.")
