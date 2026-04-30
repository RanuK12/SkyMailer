import openpyxl
import sys

file1 = r'C:\Users\emilio\Desktop\Oficina Ranuk\Ranuk-Outreach\emails_listos_todos.xlsx'
file2 = r'C:\Users\emilio\Desktop\Oficina Ranuk\Ranuk-Outreach\emails_listos_58_emails.xlsx'

def get_emails_from_wb(wb_path):
    emails = set()
    wb = openpyxl.load_workbook(wb_path, data_only=True)
    if 'Import_Instantly' in wb.sheetnames:
        ws = wb['Import_Instantly']
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            headers = [str(h).strip().lower() for h in rows[0]]
            try:
                email_idx = headers.index('email')
                for row in rows[1:]:
                    email = str(row[email_idx]).strip().lower()
                    if email and email != 'none':
                        emails.add(email)
            except ValueError:
                pass
    return emails

emails_todos = get_emails_from_wb(file1)
emails_58 = get_emails_from_wb(file2)

duplicates = emails_todos.intersection(emails_58)

print(f"Total in {file1.split('\\')[-1]}: {len(emails_todos)}")
print(f"Total in {file2.split('\\')[-1]}: {len(emails_58)}")
print(f"Duplicates found: {len(duplicates)}")
for dup in duplicates:
    print(f" - {dup}")
